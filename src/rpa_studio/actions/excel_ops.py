from __future__ import annotations
from pathlib import Path
from rpa_studio.actions.base import ActionHandler
from rpa_studio.models import ActionType, Step
from rpa_studio.engine.context import ExecutionContext

class ExcelOpenHandler(ActionHandler):
    action_type = ActionType.EXCEL_OPEN
    def execute(self, step: Step, context: ExecutionContext):
        import openpyxl
        filepath = Path(context.resolve(step.params.get("file_path", "")))
        context.add_log(f"엑셀 열기: {filepath.name}")
        wb = openpyxl.load_workbook(filepath)
        context.variables["_excel_wb"] = wb
        context.variables["_excel_path"] = str(filepath)
        return {"file": str(filepath)}

class ExcelWriteHandler(ActionHandler):
    action_type = ActionType.EXCEL_WRITE
    def execute(self, step: Step, context: ExecutionContext):
        import openpyxl
        filepath = Path(context.resolve(step.params.get("file_path", context.variables.get("_excel_path", ""))))
        sheet_name = step.params.get("sheet", "Sheet1")
        cell = step.params.get("cell", "A1")
        value = context.resolve(str(step.params.get("value", "")))
        context.add_log(f"엑셀에 쓰기: {filepath.name} [{sheet_name}]{cell} = {value[:20]}")
        if filepath.exists():
            wb = openpyxl.load_workbook(filepath)
        else:
            wb = openpyxl.Workbook()
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        ws[cell] = value
        wb.save(filepath)
        return {"cell": cell, "value": value}

class ExcelReadHandler(ActionHandler):
    action_type = ActionType.EXCEL_READ
    def execute(self, step: Step, context: ExecutionContext):
        import openpyxl
        filepath = Path(context.resolve(step.params.get("file_path", context.variables.get("_excel_path", ""))))
        sheet_name = step.params.get("sheet", "Sheet1")
        cell = step.params.get("cell", "A1")
        var_name = step.params.get("save_as", "엑셀값")
        context.add_log(f"엑셀에서 읽기: {filepath.name} [{sheet_name}]{cell}")
        wb = openpyxl.load_workbook(filepath)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        value = ws[cell].value
        context.variables[var_name] = value
        return {"cell": cell, "value": value}
